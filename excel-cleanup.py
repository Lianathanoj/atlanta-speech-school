from tkinter import Tk
from os import getcwd, path
from geocoder import google
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.filedialog import askopenfilename

root = Tk()
root.withdraw()
file_path = askopenfilename(initialdir=getcwd(), title="Select an Excel file.",
                            filetypes=[("Microsoft Excel Worksheet", ".xlsx")])
wb = load_workbook(file_path)
site_sheet = wb['Site']
cell_errors = []

# Get columns letters
longCol = get_column_letter(site_sheet.max_column + 6)
latCol = get_column_letter(site_sheet.max_column + 5)
zipCol = get_column_letter(site_sheet.max_column + 4)
stateCol = get_column_letter(site_sheet.max_column + 3)
cityCol = get_column_letter(site_sheet.max_column + 2)
addressCol = get_column_letter(site_sheet.max_column + 1)

def set_modified_titles():
    site_sheet[longCol + '1'] = "Latitude"
    site_sheet[latCol + '1'] = "Longitude"
    site_sheet[zipCol + '1'] = "Address"
    site_sheet[stateCol + '1'] = "State"
    site_sheet[cityCol + '1'] = "City"
    site_sheet[addressCol + '1'] = "Modified Site Address"

def find_address_column_letter():
    for row in site_sheet.iter_rows(min_row=0, max_row=1):
        for index, cell in enumerate(row):
            if cell.value == "Site Address":
                return get_column_letter(index + 1)

def cell_operations():
    original_address_column_letter = find_address_column_letter()
    original_column_no_header = site_sheet[original_address_column_letter][1:]
    for cellindex, cell in enumerate(original_column_no_header, start=0):
        address = cell.value
        addressArray = address.split()
        # Address cleaning and reformatting operations
        for index, word in enumerate(addressArray):
            word = word.strip(',')
            word = word.strip('.')
            addressArray[index] = word
        if addressArray[len(addressArray) - 1].upper() != "GA"\
                and addressArray[len(addressArray) - 1].upper() != "GEORGIA":
            for i, partOfAdd in enumerate(addressArray):
                if i==len(addressArray) - 1:
                    site_sheet[zipCol + str(cellindex + 2)] = partOfAdd
                elif i==len(addressArray) - 2:
                    #will be state
                    #converts "Georgia" to "GA"
                    if partOfAdd == "Georgia":
                        site_sheet[stateCol + str(cellindex + 2)] = "GA"
                    else:
                        site_sheet[stateCol + str(cellindex + 2)] = partOfAdd.upper()
                elif i==len(addressArray) - 3:
                    #will be city
                    site_sheet[cityCol + str(cellindex + 2)] = partOfAdd
            streetAd = addressArray[0]
            for word in addressArray[1:len(addressArray) - 3]:
                streetAd += " " + word
                site_sheet[addressCol + str(cellindex + 2)] = streetAd
        else:
            if addressArray[len(addressArray) - 1] == "Georgia":
                site_sheet[stateCol + str(cellindex + 2)] = "GA"
            else:
                site_sheet[stateCol + str(cellindex + 2)] = addressArray[len(addressArray) - 1].upper()
            site_sheet[cityCol + str(cellindex + 2)] = addressArray[len(addressArray) - 2]
            streetAd = addressArray[0]
            for word in addressArray[1:len(addressArray) - 3]:
                streetAd += " " + word
                site_sheet[addressCol + str(cellindex + 2)] = streetAd

        # Lat/Long Operations
        try:
            latlng = google(cell.value).latlng
            lat, long = latlng[0], latlng[1]
            site_sheet[latCol + str(cellindex + 2)] = lat
            site_sheet[longCol + str(cellindex + 2)] = long
        except ValueError:
            print('Could not read address correctly.')
            cell_errors.append((cell, cell.value))

    print("These addresses could not be converted to latitude-longitude coordinates.")
    print(cell_errors)

    file_name = path.basename(file_path)
    index = file_name.find(".xlsx")
    wb.save(file_name[:index] + '_MODIFIED' + file_name[index:])

if __name__ == '__main__':
    set_modified_titles()
    cell_operations()