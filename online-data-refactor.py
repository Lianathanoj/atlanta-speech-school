from time import sleep
from tkinter import Tk
from os import getcwd, path
from geocoder import google
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from tkinter.filedialog import askopenfilename

root = Tk()
root.withdraw()
file_path = askopenfilename(initialdir=getcwd(), title="Select an Excel file to represent online data.",
                            filetypes=[("Microsoft Excel Worksheet", ".xlsx")])
wb = load_workbook(file_path)
online_sheet = wb['Online Data']
cell_errors = []

# Get columns letters
longCol = get_column_letter(online_sheet.max_column + 2)
latCol = get_column_letter(online_sheet.max_column + 1)

def set_modified_titles():
    online_sheet[longCol + '1'] = "Longitude"
    online_sheet[latCol + '1'] = "Latitude"
    
def find_country_column_index():
    for row in online_sheet.iter_rows(min_row=0, max_row=1):
        for index, cell in enumerate(row):
            if cell.value == "Country":
                return index

def find_id_column_index():
    for row in online_sheet.iter_rows(min_row=0, max_row=1):
        for index, cell in enumerate(row):
            if cell.value == "id":
                return index

def cell_operations():
    country_column_index = find_country_column_index()
    id_column_index = find_id_column_index()

    for index, row in enumerate(online_sheet.iter_rows()):
        if index == 0:
            continue
        print(row[id_column_index].value, row[country_column_index].value)
        # try to find location from id and country name
        if not found_location(str(row[id_column_index].value) + ', '
                                  + str(row[country_column_index].value), index):
            # if id and country name fails, try finding location just by id
            if not found_location(str(row[id_column_index].value), index):
                print("Could not find location.")
    file_name = path.basename(file_path)
    index = file_name.find(".xlsx")
    wb.save(file_name[:index] + '_MODIFIED' + file_name[index:])

def found_location(location_name, index):
    num_error_checks = 5
    for i in range(1, num_error_checks + 1):
        try:
            if lat_long_conversion(location_name, index):
                return True
        except Exception:
            # try again for three times to make sure the error wasn't on Google's end
            # (e.g. internal server error, 500 status)
            if i == num_error_checks:
                print('Could not read address correctly after ' + str(num_error_checks) + ' tries.')
                cell_errors.append((location_name, index))
            else:
                print("Could not read address correctly. Trying again.")

def lat_long_conversion(location, cell_index):
    # add sleep time to not throttle api with request and to prevent query limits per second
    sleep(0.1)
    ggl = google(location)
    latlng = ggl.latlng
    lat, long = latlng[0], latlng[1]
    online_sheet[latCol + str(cell_index + 1)] = lat
    online_sheet[longCol + str(cell_index + 1)] = long
    return latlng

if __name__ == '__main__':
    set_modified_titles()
    cell_operations()

